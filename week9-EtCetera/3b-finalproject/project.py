
def main():
	# libariry for coping data and uploading/inserting it to a SQLite database 
	import openpyxl
	#This libariry is neccassry to be able to write SQLite queries
	import sqlite3

	# Reading the excel file into workBook variable, openpyxl.load_workbook('nameOfExcelFile')
	workBook = openpyxl.load_workbook('members_data.xlsx')
	# The most recently edited/active sheet in the excel file, workBook['nameOfTheSheetYou want to copy data from']
	workBook_sheet = workBook['sheet1']

	# This try block checks weather there is any error that could be raised from any of the sqlite queries ((cursor.execute()))
	try:
		
		# Connect to DB and create a cursor
		sqliteConnection = sqlite3.connect('database.db')
		cursor = sqliteConnection.cursor()
		print('DB Init')	

		create_db_tables(cursor) 


							# This upcoming section loops through each row in the excel file then copies each cell to a variable #


		# Always clear all fromating from all cells and rows of the excel sheet, to make the loop stop at the last cells which contain actual data, so you don't have to use this part:

													#########################################################################

																	# for row_num, row in enumerate(iterable, start=0)
																	# if row_num == 85 :
																		# 	break

													#########################################################################
			

		measurement_date, listOfClientsWithoutPhoneNumbers = copy_client_personal_info_from_excel_and_insert_into_db(workBook_sheet, cursor)
  		    
		copy_measurements_from_excel_and_insert(measurement_date, workBook_sheet, cursor)
    
		if listOfClientsWithoutPhoneNumbers:
			msg = create_not_added_clients_msg(listOfClientsWithoutPhoneNumbers)
			print(msg)

		# This Inserts all the values from _3BTrainersOfClients_, values as the trainers' names, and each baranch the trainer works in
		cursor.execute('''
			INSERT OR IGNORE INTO Trainer_info (Trainer_system_name, branchOfTrainer)
			SELECT Trainer_system_name, branchOfTrainer
			FROM _3BTrainersOfClients_
			WHERE Trainer_system_name IS NOT NULL AND branchOfTrainer IS NOT NULL;
			''')
			
		# This Insertes into Trainers table which holds each client_id corresponded to the trainer_id assigned to him
		cursor.execute('''
					INSERT INTO Trainers (Client_id)
					SELECT Client_id FROM Client_info;
				''')
		
		# This Update/inserts values to the column of 'Trainer_id', and this query is neccessary beacause after executing the previous command the column of 'Trainer_id' will be empty (full of NULL values), so we have to Update those values and assign to them each correct 'Trainer_id'
		cursor.execute('''
			UPDATE Trainers
				SET Trainer_id = (
				SELECT trf.Trainer_id FROM Trainer_info AS trf
				JOIN _3BTrainersOfClients_ AS trc ON trf.Trainer_system_name = trc.Trainer_system_name
				WHERE Trainers.Client_id = trc.Client_id
				)
				WHERE Client_id IN (
					SELECT Client_id 
					FROM _3BTrainersOfClients_
				);
			''')
  
		# To save the changes I have made permeneantly
		sqliteConnection.commit()

		# Close the cursor
		cursor.close()
		
	# Handle errors
	except sqlite3.Error as error:
		print('Error occurred - ', error)


	# This is a part of the try statement which doo execute wither any part of try statemnet happend
	# Close DB Connection irrespective of success
	# or failure
	finally:
		if sqliteConnection:

			sqliteConnection.close()
			print('SQLite Connection closed')	


def create_db_tables(cursor):
    
	# Sqlite commands executed by python code

					#   CREATING the 'schema'  #

	# This command Create the first table "Client_info"
	cursor.execute('''
		CREATE TABLE IF NOT EXISTS Client_info (
			Client_id INTEGER NOT NULL,
			Name TEXT,
			Phone_no TEXT UNIQUE,
			AGE TEXT,
			Height REAL,
			Weight REAL,
			Activity TEXT,
			PRIMARY KEY (Client_id)
			);''')

	
	# This command Create table named "Dyno_measurments"

	cursor.execute('''
		CREATE TABLE IF NOT EXISTS Dyno_measurments (
			Dyno_id INTEGER NOT NULL,
			Client_id INTEGER,
			Date NUMERIC, 
			Energy_src_from_protein__percent TEXT,
			Energy_src_from_carbs__percent TEXT,
			Energy_src_from_fat__percent TEXT,
			BMR_analysis_value__kcal TEXT,
			BMR_ruleOfThumb__kcal TEXT,
			BMR_calorie_requirment__kcal TEXT,
			Combustion_perDay__kcal TEXT,
			Calorie_consumption_through_activities__kcal TEXT,
			PRIMARY KEY (Dyno_id),
			FOREIGN KEY (Client_id) REFERENCES Client_info(Client_id)
			);''')
	
	# This command Create table named "Caliper_measurments"

	cursor.execute('''
		CREATE TABLE IF NOT EXISTS Caliper_measurments (
			Caliper_id INTEGER NOT NULL,
			Client_id INTEGER,
			Date NUMERIC, 
			Chest__mm TEXT,
			Scapula__mm TEXT,
			Axilla__mm TEXT,
			Triceps__mm TEXT,
			Abdominal__mm TEXT,
			Suprailiac__mm TEXT,
			Thigh__mm TEXT,
			Body_fat__percent TEXT,
			Max_preferred_level__percent TEXT,
			LBM__Lean_body_mass__kg TEXT,
			FBM__Fat_body_mass__kg TEXT,
			BMI__Body_mass_index__kg TEXT,
			BMR__Bascal_metabolic_rate__kcal TEXT,
			PRIMARY KEY (Caliper_id),
			FOREIGN KEY (Client_id) REFERENCES Client_info(Client_id)
			);''')
	
	# This command Create table named "Branch"

	cursor.execute('''
		CREATE TABLE IF NOT EXISTS Branch (
			Client_id INTEGER,
			Branch_name TEXT,
			FOREIGN KEY (Client_id) REFERENCES Client_info(Client_id)
			);''')

	# This command Create table named "Tariners_info"

	cursor.execute('''
		CREATE TABLE IF NOT EXISTS Trainer_info (
			Trainer_id INTEGER NOT NULL,
			Trainer_system_name TEXT UNIQUE,
			branchOfTrainer TEXT,
			PRIMARY KEY (Trainer_id)
			);''')
	
	# This command Create table named "Trainers"

	cursor.execute('''
		CREATE TABLE IF NOT EXISTS Trainers (
			Trainer_id INTEGER,
			Client_id INTEGER,
			FOREIGN KEY (Trainer_id) REFERENCES Trainer_info(Trainer_id),
			FOREIGN KEY (Client_id) REFERENCES Client_info(Client_id)
			PRIMARY KEY (Trainer_id, Client_id)
		);''')

	# This creats a table which is neccassry to later copy the data from it to "Trainers" table
	# ACtually you can be use this instead of Trainers, in searching and inserting etc.....
	cursor.execute('''
		CREATE TABLE IF NOT EXISTS _3BTrainersOfClients_ (
			Client_id INTEGER NOT NULL,
			Trainer_system_name TEXT,
			branchOfTrainer TEXT,
			FOREIGN KEY (Client_id) REFERENCES Client_info(Client_id)
			);''')

def copy_client_personal_info_from_excel_and_insert_into_db(workBook_sheet, cursor):
    
	# This inislizes a list that will hold values of the names of client's which there phone numbers are missing
	listOfClientsWithoutPhoneNumbers = []
  
	for row in workBook_sheet.iter_rows(min_row=5):
      
		date = row[0].value
		try:
			name = row[1].value.strip()
		except AttributeError:
			name = row[1].value

		# Also I coudnot apply .strip() the value of excel_phone, beacasue maybe the value of excel_phone is 'None' and it will raise an AttributeError if I try to do so
		try:
			excel_phone = row[2].value.strip()
			excel_phone = excel_phone.replace(' ', '')
		except AttributeError:
			excel_phone = row[2].value

		# This adds the name of the client if he did not have a phone number to a the list of people without phone numbers
		if excel_phone == None:
			listOfClientsWithoutPhoneNumbers.append(name)
		
		age = row[3].value
		height = row[4].value
		weight = row[5].value

		try:
			activity = row[6].value.strip()
		except AttributeError:
			activity = row[6].value

		# This can skip duplicated values as entries and continue with the quries
		cursor.execute('''
			INSERT OR IGNORE INTO Client_info (Name, Phone_no, AGE, Height, Weight, Activity)
			VALUES (?, ?, ?, ?, ?, ?)
			''', (name, excel_phone, age, height, weight, activity)
			)
  
	return date, listOfClientsWithoutPhoneNumbers


def extract_client_phone_numbers_in_db(cursor):
    # This block of code fetches all phone numbers in the data base, including the the one that recently got added
	cursor.execute('''
				SELECT Phone_no FROM Client_info;
		''')
	# 'listOfPhoneNumbersIn_db' this is a -------list of tuples[(tuple one),(tuple two),(tuple three)]------, and each tuple contain one phone number-----phoneNumber_tuple =(01099133377,)
	return cursor.fetchall()

def copy_measurements_from_excel_and_insert(measurement_date, workBook_sheet, cursor):
    
	for row in workBook_sheet.iter_rows(min_row=5):
		# Read again each phone number in each row in the excel file found in column (phone number)

		try:
			excel_phone = ''.join(row[2].value.strip())
		except AttributeError:
			excel_phone = str(row[2].value).strip()


		# This part copies the Branch & Trainer_info info cells in from the (Branch & Trainer_info) columns found in excel sheet 
		try:
			Trainer_system_name = row[7].value.strip()
		except AttributeError:
			Trainer_system_name = row[7].value

		try:
			Branch_name = row[8].value.strip()
		except AttributeError:
			Branch_name = row[8].value

		try:
			branchOfTrainer = row[8].value.strip()
		except AttributeError:
			branchOfTrainer = row[8].value

		# This part copies the Measuremnts of the dynostics devices in from the (Dynostics Measuremnts) columns found in excel sheet 

		try:
			Energy_src_from_protein__percent = row[9].value.strip()
		except AttributeError:
			Energy_src_from_protein__percent = row[9].value
		
		try:
			Energy_src_from_carbs__percent = row[10].value.strip()
		except AttributeError:
			Energy_src_from_carbs__percent = row[10].value
		
		try:
			Energy_src_from_fat__percent = row[11].value.strip()
		except AttributeError:
			Energy_src_from_fat__percent = row[11].value
		
		try:
			BMR_analysis_value__kcal = row[12].value.strip()
		except AttributeError:
			BMR_analysis_value__kcal = row[12].value
		
		try:
			BMR_ruleOfThumb__kcal = row[13].value.strip()
		except AttributeError:
			BMR_ruleOfThumb__kcal = row[13].value
		
		try:
			BMR_calorie_requirment__kcal = row[14].value.strip()
		except AttributeError:
			BMR_calorie_requirment__kcal = row[14].value

		try:
			Combustion_perDay__kcal = row[15].value.strip()
		except AttributeError:
			Combustion_perDay__kcal = row[15].value
		
		try:
			Calorie_consumption_through_activities__kcal = row[16].value.strip()
		except AttributeError:
			Calorie_consumption_through_activities__kcal = row[16].value
		

		# This part copies the Measuremnts of the caliper tool in from the (Calliper Measurements) columns found in excel sheet 


		try:
			Chest__mm = row[17].value.strip()
		except AttributeError:
			Chest__mm = row[17].value

		try:
			Scapula__mm = row[17].value.strip()
		except AttributeError:
			Scapula__mm = row[17].value
		
		try:
			Axilla__mm = row[19].value.strip()
		except AttributeError:
			Axilla__mm = row[19].value
		
		try:
			Triceps__mm = row[20].value.strip()
		except AttributeError:
			Triceps__mm = row[20].value

		try:
			Abdominal__mm = row[21].value.strip()
		except AttributeError:
			Abdominal__mm = row[21].value
		
		try:
			Suprailiac__mm = row[22].value.strip()
		except AttributeError:
			Suprailiac__mm = row[22].value
		
		try:
			Thigh__mm = row[23].value.strip()
		except AttributeError:
			Thigh__mm = row[23].value
		
		try:
			Body_fat__percent = row[24].value.strip()
		except AttributeError:
			Body_fat__percent = row[24].value

		try:
			Max_preferred_level__percent = row[25].value.strip()
		except AttributeError:
			Max_preferred_level__percent = row[25].value

		try:
			LBM__Lean_body_mass__kg = row[26].value.strip()
		except AttributeError:
			LBM__Lean_body_mass__kg = row[26].value

		try:
			FBM__Fat_body_mass__kg = row[27].value.strip()
		except AttributeError:
			FBM__Fat_body_mass__kg = row[27].value

		try:
			BMI__Body_mass_index__kg = row[28].value.strip()
		except AttributeError:
			BMI__Body_mass_index__kg = row[28].value

		try:
			BMR__Bascal_metabolic_rate__kcal = row[29].value.strip()
		except AttributeError:
			BMR__Bascal_metabolic_rate__kcal = row[29].value
	
		# Used for checking wheather the upcoming part of code got triggerd or not, as if it actually didnot get triggerd that would mean that the client is not in the data base (hopefully means that there is no phone number written in the excel sheet)
		clientHasBeenAddedToList = False
  
		listOfPhoneNumbersIn_db = extract_client_phone_numbers_in_db(cursor)
  
		for phoneNumber_tuple in listOfPhoneNumbersIn_db:
			for phoneNumberIn_db in phoneNumber_tuple:
				if phoneNumberIn_db == excel_phone:

														#+$#_#+_$+_#+$_$+_#+_#+_#$+#_$+#_$#+_#$+_#$+#_$+#_$+#$_
					# I don't know why this didn't work (doesnot work in the first place)
					# cursor.execute('''
					# 				SELECT Client_id FROM Client_info
					# 				WHERE phone_no = ?
					# 			''', (excel_phone,))
					
					# I dont know why the F&%$ the upcoming line doesnot work (Assigns a None value to Cleint_id ?????)
					# cursor.execute('SELECT Client_id FROM Client_info WHERE Phone_no = "{}"'.format("excel_phone"))
														#+$#_#+_$+_#+$_$+_#+_#+_#$+#_$+#_$#+_#$+_#$+#_$+#_$+#$_


					# This fetches the client's id from the data base based on his UNIQUE personal number
					query = f"SELECT Client_id FROM Client_info WHERE Phone_no = '{excel_phone}'"
					cursor.execute(query)

					Client_id_tuple = cursor.fetchone()
					Client_id = Client_id_tuple[0]

					# This copies the Dyno measurments columns from sheet to our db
					cursor.execute('''
						INSERT INTO Dyno_measurments (Client_id, Date, Energy_src_from_protein__percent, Energy_src_from_carbs__percent, Energy_src_from_fat__percent, BMR_analysis_value__kcal, BMR_ruleOfThumb__kcal, BMR_calorie_requirment__kcal, Combustion_perDay__kcal, Calorie_consumption_through_activities__kcal)
						VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
					''', (Client_id, measurement_date, Energy_src_from_protein__percent, Energy_src_from_carbs__percent, Energy_src_from_fat__percent, BMR_analysis_value__kcal, BMR_ruleOfThumb__kcal, BMR_calorie_requirment__kcal, Combustion_perDay__kcal, Calorie_consumption_through_activities__kcal)
					)

					# This copies the Caliper measurments columns from sheet to our db
					cursor.execute('''
						INSERT INTO Caliper_measurments (Client_id, Date, Chest__mm, Scapula__mm, Axilla__mm, Triceps__mm, Abdominal__mm, Suprailiac__mm, Thigh__mm, Body_fat__percent, Max_preferred_level__percent, LBM__Lean_body_mass__kg, FBM__Fat_body_mass__kg, BMI__Body_mass_index__kg, BMR__Bascal_metabolic_rate__kcal)
						VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
						''', (Client_id, measurement_date, Chest__mm, Scapula__mm, Axilla__mm, Triceps__mm, Abdominal__mm, Suprailiac__mm, Thigh__mm, Body_fat__percent, Max_preferred_level__percent, LBM__Lean_body_mass__kg, FBM__Fat_body_mass__kg, BMI__Body_mass_index__kg, BMR__Bascal_metabolic_rate__kcal)
						)
					
					# This copies the Branch data fcolumns from sheet to our db
					cursor.execute('''
						INSERT INTO Branch (Client_id, Branch_name)
						VALUES (?, ?)
						''', (Client_id, Branch_name)
						)
					
					# This copies the Trainer_info columns from sheet to our db
					cursor.execute('''
						INSERT INTO _3BTrainersOfClients_ (Client_id, Trainer_system_name, branchOfTrainer)
						VALUES (?, ?, ?)
						''', (Client_id, Trainer_system_name, branchOfTrainer)
						)
					
					clientHasBeenAddedToList = True
					# This will break the inner loop, after the previous part (hpefully) executed sccessfully, and now I want to move to the next value of excel_phone from a new row in the excel sheet
					break


			# This will break the outer loop 
			if clientHasBeenAddedToList:
				break
			clientHasBeenAddedToList = False			

def create_not_added_clients_msg(listOfClientsWithoutPhoneNumbers):
	return f"****\nNotice That clients listed in {listOfClientsWithoutPhoneNumbers} do not have a phone number provided!\n****"

main()